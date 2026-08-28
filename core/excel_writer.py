"""Módulo de modificación y combinación en Excel.

Cambios v2:
  - Nuevo: rellenar_formulario_excel() retorna (bytes, reporte) donde reporte
    es una lista de dicts con el estado de cada campo (escrito / saltado / error).
  - Fix: ubicacion "abajo" ahora calcula correctamente la columna de destino
    cuando el rótulo origen está en un rango merge (usa min_col del merge).
  - Fix: _obtener_celda_escribible() maneja MergedCell en dirección vertical.
  - Nuevo: _log_item() centraliza el log por campo para diagnóstico.
"""

from copy import copy
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Helpers de celda
# ---------------------------------------------------------------------------

def _celda_en_merge(hoja: Worksheet, fila: int, columna: int) -> Optional[Any]:
    for rango in hoja.merged_cells.ranges:
        if (rango.min_row <= fila <= rango.max_row
                and rango.min_col <= columna <= rango.max_col):
            return rango
    return None


def _calcular_max_columna_real(ws: Worksheet) -> int:
    """Calcula el índice máximo de columna que posee contenido o bordes reales en el formulario,
    evitando que openpyxl reporte un ws.max_column inflado por columnas vacías con estilos a la derecha.
    """
    max_c = 1
    max_filas = min(ws.max_row or 1, 150)
    max_cols_eval = min(ws.max_column or 1, 80)

    for r in range(1, max_filas + 1):
        for c in range(1, max_cols_eval + 1):
            celda = ws.cell(row=r, column=c)
            if celda.value is not None and str(celda.value).strip():
                if c > max_c:
                    max_c = c
            elif celda.border:
                b = celda.border
                if ((b.left and b.left.style and b.left.style != "none") or
                    (b.right and b.right.style and b.right.style != "none") or
                    (b.top and b.top.style and b.top.style != "none") or
                    (b.bottom and b.bottom.style and b.bottom.style != "none")):
                    if c > max_c:
                        max_c = c

    return max(max_c, 1)


def _buscar_campo_anidado(obj: Any, campo: str) -> Any:

    if isinstance(obj, dict):
        if campo in obj:
            return obj[campo]
        for value in obj.values():
            resultado = _buscar_campo_anidado(value, campo)
            if resultado is not None:
                return resultado
    return None


def _obtener_valor_datos(datos_empresa: Dict[str, Any], campo: str) -> Any:
    """Resuelve el valor del campo, incluyendo campos virtuales nit_sin_dv / nit_dv y taxonomía jerárquica."""
    from core.profile_manager import aplanar_perfil
    plano = aplanar_perfil(datos_empresa)

    if campo in ("ciudad_departamento", "ciudad/departamento", "ciudad_depto"):
        c = str(plano.get("ciudad", "")).strip()
        d = str(plano.get("departamento", "")).strip()
        if c and d:
            return f"{c}/{d}"
        return c or d

    if campo == "representante_nombres":
        val = plano.get("representante_nombres")
        if val and str(val).strip():
            return val
        rep_full = str(plano.get("representante_legal", "")).strip()
        if rep_full:
            partes = rep_full.split()
            return " ".join(partes[:-2]) if len(partes) > 2 else (partes[0] if partes else rep_full)

    if campo == "representante_apellidos":
        val = plano.get("representante_apellidos")
        if val and str(val).strip():
            return val
        rep_full = str(plano.get("representante_legal", "")).strip()
        if rep_full:
            partes = rep_full.split()
            return " ".join(partes[-2:]) if len(partes) >= 2 else ""

    if campo in plano:
        val = plano[campo]
        if not isinstance(val, dict):
            return val

    if campo == "identificacion" and "identificacion" not in plano:
        return plano.get("cedula")

    # Acceso por ruta anidada terminal
    if "." in campo:
        subcampo = campo.split(".")[-1]
        if subcampo in plano and not isinstance(plano[subcampo], dict):
            return plano[subcampo]

    return None


def _obtener_relleno_preservado(celda_origen, celda_destino) -> Optional[PatternFill]:
    """WRITER-01: Preserva el fondo original de la celda destino únicamente.
    NUNCA copia el color de fondo del rótulo/encabezado (celda_origen) hacia el área de escritura.
    """
    if celda_destino and hasattr(celda_destino, "fill") and celda_destino.fill:
        fill_type = getattr(celda_destino.fill, "fill_type", None)
        if fill_type is not None and fill_type != "none":
            return copy(celda_destino.fill)
    return None


def _obtener_fuente_preservada(celda_origen, celda_destino) -> Optional[Font]:
    """WRITER-02: Preserva la tipografía (nombre de fuente, tamaño) adaptada para escritura.
    Hereda la familia y tamaño de fuente del rótulo origen o celda destino, manteniendo
    el texto legible en estilo normal (no negrita decorativa de encabezado).
    """
    # 1. Heredar familia y tamaño del rótulo origen (con bold=False para texto de entrada limpio)
    if celda_origen and hasattr(celda_origen, "font") and celda_origen.font:
        f_orig = celda_origen.font
        if f_orig.name or f_orig.size:
            return Font(
                name=f_orig.name or "Calibri",
                size=f_orig.size or 10,
                bold=False,
                italic=bool(f_orig.italic),
                color=copy(f_orig.color) if f_orig.color else None,
            )

    # 2. Si la celda destino tiene fuente propia definida
    if celda_destino and hasattr(celda_destino, "font") and celda_destino.font:
        f_dest = celda_destino.font
        return Font(
            name=f_dest.name or "Calibri",
            size=f_dest.size or 10,
            bold=False,
            italic=bool(f_dest.italic),
            color=copy(f_dest.color) if f_dest.color else None,
        )

    return Font(name="Calibri", size=10, bold=False)


def _obtener_borde_preservado(celda_origen, celda_destino, es_misma: bool = False) -> Optional[Border]:
    """WRITER-04: Preserva los bordes de la celda de destino y evita crear líneas verticales parásitas a la izquierda.

    Reglas:
      1. Si la celda destino tiene bordes propios definidos, se preservan fielmente.
      2. Si se escribe en la misma celda (es_misma=True), preserva el borde original.
      3. Si se escribe en una celda adyacente (derecha / abajo):
         - NUNCA se copia el borde 'left' ni 'right' de celda_origen (esto creaba un borde vertical parásito al inicio de la palabra).
         - Solo se hereda el borde 'bottom' (subrayado) si celda_origen lo tenía y celda_destino no tiene borde inferior.
    """
    b_dest = celda_destino.border if celda_destino and hasattr(celda_destino, "border") else None
    b_orig = celda_origen.border if celda_origen and hasattr(celda_origen, "border") else None

    if b_dest:
        tiene_bordes_dest = any(
            getattr(b_dest, l, None) and getattr(b_dest, l).style and getattr(b_dest, l).style != "none"
            for l in ("top", "bottom", "left", "right")
        )
        if tiene_bordes_dest:
            return copy(b_dest)

    if es_misma and b_orig:
        return copy(b_orig)

    # Escritura adyacente: solo heredar subrayado inferior si la celda origen lo tenía
    if b_orig and b_orig.bottom and b_orig.bottom.style and b_orig.bottom.style != "none":
        return Border(bottom=copy(b_orig.bottom))

    return None


def _obtener_celda_escribible(hoja: Worksheet, fila: int, columna: int) -> Any:
    """Retorna siempre un objeto Cell escribible (nunca MergedCell)."""
    for rango in hoja.merged_cells.ranges:
        if (rango.min_row <= fila <= rango.max_row
                and rango.min_col <= columna <= rango.max_col):
            return hoja.cell(row=rango.min_row, column=rango.min_col)
    celda = hoja.cell(row=fila, column=columna)
    if type(celda).__name__ == "MergedCell":
        for rango in hoja.merged_cells.ranges:
            if (rango.min_row <= fila <= rango.max_row
                    and rango.min_col <= columna <= rango.max_col):
                return hoja.cell(row=rango.min_row, column=rango.min_col)
    return celda


def _escribir_valor_en_celda(celda, valor: Any, es_misma_celda: bool, hoja: Optional[Worksheet] = None) -> bool:
    """Escribe el valor en la celda. Retorna True si se escribió, False si se saltó."""
    if type(celda).__name__ == "MergedCell":
        if hoja is not None:
            celda = _obtener_celda_escribible(hoja, celda.row, celda.column)
            if type(celda).__name__ == "MergedCell":
                return False
        else:
            return False

    if isinstance(valor, bool):
        valor = "X" if valor else ""

    valor_actual = celda.value

    # Limpiar valor actual de caracteres invisibles o de formato
    txt_actual = ""
    if valor_actual is not None:
        txt_actual = str(valor_actual).replace("\xa0", " ").replace("\t", " ").strip()

    patron_placeholder = r'_{2,}|\.{3,}'

    # Si la celda contiene una fórmula o comilla simple que el usuario ve vacía en Excel, permitir sobreescritura
    es_formula_o_vacio_visualmente = bool(
        not txt_actual or
        txt_actual.startswith("=") or
        txt_actual in ("''", '""', "-", "N/A", "0") or
        re.match(r"^[\s_\.\:\-]+$", txt_actual)
    )

    # CASO A: Escritura en celda adyacente (derecha / abajo)
    if not es_misma_celda:
        if not es_formula_o_vacio_visualmente:
            if re.search(patron_placeholder, txt_actual):
                try:
                    celda.value = re.sub(patron_placeholder, str(valor), txt_actual, count=1)
                    return True
                except AttributeError:
                    return False
            # Celda con contenido descriptivo real pre-impreso: no sobreescribir
            return False

        try:
            celda.value = valor
            return True
        except AttributeError:
            return False

    # CASO B: Escritura en la MISMA celda (dirección = "misma")
    if es_formula_o_vacio_visualmente:
        # Celda estaba vacía o solo contenía líneas/guiones puros
        try:
            celda.value = valor
            return True
        except AttributeError:
            return False

    # 1. Si contiene líneas/guiones inline combinados con texto (ej: "NIT: ________" o "Yo, ________"):
    if re.search(patron_placeholder, txt_actual):
        try:
            celda.value = re.sub(patron_placeholder, str(valor), txt_actual, count=1)
            return True
        except AttributeError:
            return False

    # 2. Si termina en dos puntos (ej: "NOMBRE / RAZON SOCIAL:" o "DIRECCIÓN:")
    if txt_actual.endswith(":"):
        try:
            celda.value = f"{txt_actual} {valor}"
            return True
        except AttributeError:
            return False

    # 3. Si contiene dos puntos en el texto (ej: "Ciudad : ")
    if ":" in txt_actual:
        partes = txt_actual.split(":", 1)
        rotulo_limpio = partes[0].strip()
        try:
            celda.value = f"{rotulo_limpio}: {valor}"
            return True
        except AttributeError:
            return False

    # 4. Si es un rótulo sin dos puntos (ej: "NOMBRE / RAZON SOCIAL"):
    try:
        celda.value = f"{txt_actual}: {valor}"
        return True
    except AttributeError:
        return False



# ---------------------------------------------------------------------------
# Log por ítem
# ---------------------------------------------------------------------------

def _log_item(
    estado: str,
    item: Dict[str, Any],
    valor: Any,
    fila_dest: int,
    col_dest: int,
    motivo: str = "",
) -> Dict[str, Any]:
    """Centraliza el log por campo y retorna un registro de reporte."""
    icono = {"OK": "✅", "SKIP": "⏭️", "ERROR": "❌", "NULL": "🔕"}.get(estado, "❓")
    msg = (
        f"[AutoForm Writer] {icono} [{estado}] "
        f"campo='{item.get('campo')}' valor='{str(valor)[:40]}' "
        f"→ {item.get('hoja')}!R{fila_dest}C{col_dest}"
    )
    if motivo:
        msg += f" ({motivo})"
    try:
        print(msg)
    except UnicodeEncodeError:
        print(msg.encode("ascii", errors="replace").decode("ascii"))
    return {
        "estado": estado,
        "campo": item.get("campo", ""),
        "valor_intentado": str(valor)[:80] if valor is not None else None,
        "hoja": item.get("hoja", ""),
        "fila_destino": fila_dest,
        "columna_destino": col_dest,
        "motivo": motivo,
    }


# ---------------------------------------------------------------------------
# Función principal de escritura
# ---------------------------------------------------------------------------

def rellenar_formulario_excel(
    bytes_excel: bytes,
    plan_mapeo: List[Dict[str, Any]],
    datos_empresa: Dict[str, Any],
    celdas_prellenadas: Optional[Dict] = None,
) -> Tuple[bytes, List[Dict[str, Any]]]:
    """Escribe el plan de mapeo en el Excel conservando estilos originales.

    Soporta Reprocesamiento Incremental mediante el parámetro `celdas_prellenadas`:
    Si se provee, las coordenadas que ya contienen el valor correcto se marcan como
    PRESERVED sin tocar el archivo, garantizando que ningún dato ya diligenciado
    se pierda o sobreescriba incorrectamente.

    Returns:
        Tuple[bytes_excel_modificado, reporte_de_inyeccion]
        El reporte contiene una entrada por ítem con estado OK/SKIP/NULL/ERROR/PRESERVED.
    """
    workbook = load_workbook(filename=BytesIO(bytes_excel), data_only=False)
    reporte: List[Dict[str, Any]] = []
    _celdas_pre = celdas_prellenadas or {}

    for item in plan_mapeo:
        hoja_nombre = str(item.get("hoja", ""))
        if hoja_nombre not in workbook.sheetnames:
            reporte.append(_log_item("ERROR", item, None, 0, 0, f"Hoja '{hoja_nombre}' no existe"))
            continue

        ws = workbook[hoja_nombre]
        fila_origen    = int(item.get("fila", 0))
        columna_origen = int(item.get("columna", 0))
        ubicacion      = str(item.get("ubicacion", "")).lower()
        rango_origen   = _celda_en_merge(ws, fila_origen, columna_origen)

        # ── Calcular coordenadas de destino ───────────────────────────────
        if ubicacion == "misma":
            fila_destino    = fila_origen
            columna_destino = columna_origen

        elif ubicacion == "derecha":
            max_col_real = _calcular_max_columna_real(ws)
            max_col_origen = rango_origen.max_col if rango_origen is not None else columna_origen

            # Si la columna destino a la derecha excede la columna máxima del formulario, escribir ABAJO
            col_dest_prevista = (rango_origen.max_col + 1) if rango_origen is not None else (columna_origen + 1)
            es_borde_derecho = (col_dest_prevista > max_col_real)
            es_merge_ancho = (rango_origen is not None and (rango_origen.max_col - rango_origen.min_col + 1) >= 15)

            if es_borde_derecho or es_merge_ancho:
                ubicacion = "abajo"
                if rango_origen is not None:
                    fila_destino = rango_origen.max_row + 1
                    columna_destino = rango_origen.min_col
                else:
                    fila_destino = fila_origen + 1
                    columna_destino = columna_origen
            else:
                fila_destino = fila_origen
                if rango_origen is not None:
                    columna_destino = rango_origen.max_col + 1
                else:
                    columna_destino = columna_origen + 1



        elif ubicacion == "abajo":
            if rango_origen is not None:
                fila_destino    = rango_origen.max_row + 1
                columna_destino = rango_origen.min_col
            else:
                fila_destino    = fila_origen + 1
                columna_destino = columna_origen

            # Garantizar que fila_destino sea estrictamente posterior al rango de la cabecera origen
            rango_dest_chk = _celda_en_merge(ws, fila_destino, columna_destino)
            if rango_dest_chk is not None and (rango_dest_chk.min_row <= fila_origen <= rango_dest_chk.max_row):
                fila_destino = rango_dest_chk.max_row + 1
        else:

            reporte.append(_log_item("ERROR", item, None, 0, 0, f"Ubicación inválida: '{ubicacion}'"))
            continue

        # ── WRITER-07: Validación de coordenadas fuera de rango ──────────
        max_fila = ws.max_row or 0
        max_col  = ws.max_column or 0
        # FIX-5: openpyxl puede subestimar max_row/max_col cuando las últimas
        # filas/columnas solo tienen bordes sin contenido. Toleramos un margen de
        # +10 filas y +5 columnas para no descartar campos legítimos al borde del formulario.
        if (fila_destino < 1 or columna_destino < 1
                or fila_destino > max_fila + 10 or columna_destino > max_col + 5):
            reporte.append(_log_item(
                "SKIP", item, None, fila_destino, columna_destino,
                f"Coordenada fuera de rango (max_fila={max_fila}, max_col={max_col})"
            ))
            continue

        # ── REPROCESAMIENTO INCREMENTAL: Verificar si la celda ya tiene el valor correcto ──
        valor_esperado = str(_obtener_valor_datos(datos_empresa, str(item.get("campo", ""))) or "").strip()
        clave_celda = (hoja_nombre, fila_destino, columna_destino)
        if _celdas_pre and clave_celda in _celdas_pre:
            val_existente = str(_celdas_pre[clave_celda]).strip()
            if val_existente and val_existente == valor_esperado:
                reporte.append(_log_item("PRESERVED", item, valor_esperado, fila_destino, columna_destino,
                                         "Valor ya diligenciado correctamente en ejecución anterior"))
                continue

        # ── WRITER-03: Determinar cantidad de columnas a combinar ─────────
        requiere_merge  = bool(item.get("requiereMerge", False))
        celdas_a_mergear = int(item.get("celdasAMergear", 1) or 1)
        ancho_linea      = int(item.get("anchoLinea", 1) or 1)

        cant_cols_merge = 1
        if requiere_merge and celdas_a_mergear > 1:
            cant_cols_merge = celdas_a_mergear
        if ancho_linea > 1:
            cant_cols_merge = max(cant_cols_merge, ancho_linea)

        # Protección: no invadir rótulos vecinos con contenido ni rangos combinados adyacentes
        if cant_cols_merge > 1:
            max_cols_libres = 1
            for col_chk in range(
                columna_destino + 1,
                min(columna_destino + cant_cols_merge, max_col + 1),
            ):
                val_chk = ws.cell(row=fila_destino, column=col_chk).value
                rango_chk = _celda_en_merge(ws, fila_destino, col_chk)
                # Detener expansión si hay contenido o si inicia un rango combinado preexistente
                if (val_chk is not None and str(val_chk).strip() != "") or (rango_chk is not None and rango_chk.min_col == col_chk):
                    break
                max_cols_libres += 1
            cant_cols_merge = max_cols_libres

        # ── Obtener el valor del campo ────────────────────────────────────
        campo = str(item.get("campo", ""))
        valor = _obtener_valor_datos(datos_empresa, campo)

        # WRITER-05: Skip silencioso si el valor es None
        if valor is None:
            reporte.append(_log_item(
                "NULL", item, None, fila_destino, columna_destino,
                f"Campo '{campo}' no encontrado en DatosEmpresa"
            ))
            continue

        # ── Aplicar merge si corresponde ──────────────────────────────────
        rango_preexistente = _celda_en_merge(ws, fila_destino, columna_destino)
        rango_combinado: Optional[Tuple[int, int, int, int]] = None

        if (cant_cols_merge > 1 and ubicacion == "derecha"
                and rango_preexistente is None):
            col_final = columna_destino + cant_cols_merge - 1
            rango_combinado = (fila_destino, columna_destino, fila_destino, col_final)
            try:
                ws.merge_cells(
                    start_row=fila_destino, start_column=columna_destino,
                    end_row=fila_destino,   end_column=col_final,
                )
            except Exception as exc_merge:
                reporte.append(_log_item(
                    "ERROR", item, valor, fila_destino, columna_destino,
                    f"Error al combinar celdas: {exc_merge}"
                ))
                continue

        # ── Obtener celdas y escribir ─────────────────────────────────────
        celda_destino = _obtener_celda_escribible(ws, fila_destino, columna_destino)
        celda_origen  = _obtener_celda_escribible(ws, fila_origen,  columna_origen)
        es_misma      = (celda_destino.coordinate == celda_origen.coordinate)

        escrito = _escribir_valor_en_celda(celda_destino, valor, es_misma, hoja=ws)

        # ── FALLBACK AUTOMÁTICO A ABAJO SI DERECHA ESTÁ BLOQUEADA O FUERA DE LÍMITE ──
        if (not escrito or columna_destino >= max_col) and ubicacion == "derecha":
            if rango_origen is not None:
                fb_fila = rango_origen.max_row + 1
                fb_col  = rango_origen.min_col
            else:
                fb_fila = fila_origen + 1
                fb_col  = columna_origen

            if 1 <= fb_fila <= max_fila and 1 <= fb_col <= max_col:
                celda_fb = _obtener_celda_escribible(ws, fb_fila, fb_col)
                if _escribir_valor_en_celda(celda_fb, valor, False, hoja=ws):
                    fila_destino = fb_fila
                    columna_destino = fb_col
                    celda_destino = celda_fb
                    escrito = True
                    print(f"[AutoForm Writer Fallback] Campo '{campo}' ({valor}) re-enrutado a ABAJO R{fb_fila}C{fb_col} al estar DERECHA ocupada o en el límite del formulario.")



        # ── WRITER-01/02/04: Preservar estilos ───────────────────────────
        relleno         = _obtener_relleno_preservado(celda_origen, celda_destino)
        fuente          = _obtener_fuente_preservada(celda_origen, celda_destino)
        borde_preservado = _obtener_borde_preservado(celda_origen, celda_destino, es_misma=es_misma)

        if rango_combinado is not None:
            _, ini_col, _, fin_col = rango_combinado
            for col in range(ini_col, fin_col + 1):
                c = _obtener_celda_escribible(ws, fila_destino, col)
                if type(c).__name__ != "MergedCell":
                    if c.alignment:
                        c.alignment = copy(c.alignment)
                    else:
                        c.alignment = Alignment(vertical="center")
                    if relleno:
                        c.fill = copy(relleno)
                    if fuente:
                        c.font = copy(fuente)
                    if borde_preservado:
                        c.border = Border(
                            top=copy(borde_preservado.top) if borde_preservado.top else None,
                            bottom=copy(borde_preservado.bottom) if borde_preservado.bottom else None,
                            left=copy(borde_preservado.left) if col == ini_col and borde_preservado.left else None,
                            right=copy(borde_preservado.right) if col == fin_col and borde_preservado.right else None,
                        )
        else:
            if type(celda_destino).__name__ != "MergedCell":
                if celda_destino.alignment:
                    celda_destino.alignment = copy(celda_destino.alignment)
                else:
                    celda_destino.alignment = Alignment(vertical="center")
                if relleno:
                    celda_destino.fill = relleno
                if fuente:
                    celda_destino.font = fuente
                if borde_preservado:
                    celda_destino.border = borde_preservado


        # ── READ-BACK VERIFICATION: Comprobación física real en el Excel ───
        verificado_real = False
        if escrito:
            celda_chk = _obtener_celda_escribible(ws, fila_destino, columna_destino)
            val_fisico = celda_chk.value
            if val_fisico is not None and str(val_fisico).strip() != "":
                verificado_real = True

        estado = "OK" if verificado_real else "SKIP"
        motivo = "" if verificado_real else ("Valor no verificado físicamente en la celda destino" if escrito else "Celda no escribible o ya contiene contenido")
        reporte.append(_log_item(estado, item, valor, fila_destino, columna_destino, motivo))

    # ── Serializar y retornar ─────────────────────────────────────────────
    salida = BytesIO()
    workbook.save(salida)
    salida.seek(0)


    # Resumen final en consola
    ok_count        = sum(1 for r in reporte if r["estado"] == "OK")
    skip_count      = sum(1 for r in reporte if r["estado"] == "SKIP")
    null_count      = sum(1 for r in reporte if r["estado"] == "NULL")
    err_count       = sum(1 for r in reporte if r["estado"] == "ERROR")
    preserved_count = sum(1 for r in reporte if r["estado"] == "PRESERVED")
    # Formatear resumen con emojis solicitados, sin PRESERVED
    summary_msg = (
        f"\n📊 Resumen: ✅ OK={ok_count}  ⏭️ SKIP={skip_count}  🔕 NULL={null_count}  ❌ ERROR={err_count}\n"
    )
    try:
        print(summary_msg)
    except UnicodeEncodeError:
        print(summary_msg.encode("ascii", errors="replace").decode("ascii"))

    return salida.getvalue(), reporte