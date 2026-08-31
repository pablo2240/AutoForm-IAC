"""Módulo de lectura de celdas y vecinos (openpyxl)."""

from __future__ import annotations

import io
from typing import Any, BinaryIO, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def cargar_libro(ruta_o_buffer: Any):
    """Carga un libro de Excel conservando el formato original."""
    if hasattr(ruta_o_buffer, "read"):
        ruta_o_buffer.seek(0)
        return load_workbook(ruta_o_buffer, data_only=False)
    return load_workbook(ruta_o_buffer, data_only=False)


def leer_celda(hoja: Worksheet, fila: int, col: int) -> Any:
    """Devuelve el valor de una celda (sin recalcular fórmulas)."""
    return hoja.cell(row=fila, column=col).value


def _es_celda_en_merge(hoja: Worksheet, fila: int, col: int) -> bool:
    """Determina si una celda pertenece a un rango combinado (fallback)."""
    for rango in hoja.merged_cells.ranges:
        if rango.min_row <= fila <= rango.max_row and rango.min_col <= col <= rango.max_col:
            return True
    return False


def _celda_vacia(hoja: Worksheet, fila: int, col: int) -> bool:
    """Verifica si una celda está vacía o no existe en la hoja."""
    if fila < 1 or col < 1:
        return True
    
    # Intentamos obtener la celda sin instanciarla si está fuera de las dimensiones existentes
    if fila > (hoja.max_row or 0) or col > (hoja.max_column or 0):
        return True
        
    celda = hoja.cell(row=fila, column=col)
    return celda.value is None or str(celda.value).strip() == ""


# ──────────────────────────────────────────────────────────────────────────────
# PARSER-01: Detección de bordes visuales de celdas
# ──────────────────────────────────────────────────────────────────────────────

def _tiene_borde(celda, lado: str) -> bool:
    """Devuelve True si la celda tiene un borde definido en el lado indicado.

    Args:
        celda: Objeto celda de openpyxl.
        lado: Uno de 'top', 'bottom', 'left', 'right'.
    """
    try:
        borde_lado = getattr(celda.border, lado, None)
        return borde_lado is not None and borde_lado.style is not None and borde_lado.style != "none"
    except Exception:
        return False


def _analizar_bordes_celda(hoja: Worksheet, fila: int, col: int) -> Dict[str, bool]:
    """Analiza los cuatro bordes de una celda y devuelve un diccionario con sus estados.

    Returns:
        Dict con claves: 'top', 'bottom', 'left', 'right'.
    """
    if fila < 1 or col < 1 or fila > (hoja.max_row or 0) or col > (hoja.max_column or 0):
        return {"top": False, "bottom": False, "left": False, "right": False}
    celda = hoja.cell(row=fila, column=col)
    return {
        "top":    _tiene_borde(celda, "top"),
        "bottom": _tiene_borde(celda, "bottom"),
        "left":   _tiene_borde(celda, "left"),
        "right":  _tiene_borde(celda, "right"),
    }


def _calcular_tipo_espacio(
    vacia: bool,
    es_merge: bool,
    bordes: Dict[str, bool],
) -> str:
    """Clasifica el tipo de espacio disponible para escritura basado en estado visual.

    Jerarquía de clasificación:
      - "merge"     → celda pertenece a un rango combinado (independiente del contenido)
      - "subrayado" → celda vacía con borde inferior (línea de captura _____)
      - "cuadro"    → celda vacía con todos sus bordes (caja de tabla)
      - "vacio"     → celda vacía sin bordes relevantes
      - "ocupado"   → celda no vacía (ya tiene contenido, no apta para escritura)

    Returns:
        str: "merge" | "subrayado" | "cuadro" | "vacio" | "ocupado"
    """
    if es_merge:
        return "merge"
    if not vacia:
        return "ocupado"
    tiene_bottom = bordes.get("bottom", False)
    tiene_top    = bordes.get("top", False)
    tiene_left   = bordes.get("left", False)
    tiene_right  = bordes.get("right", False)
    # Cuadro cerrado: al menos 3 de los 4 lados tienen borde
    lados_con_borde = sum([tiene_bottom, tiene_top, tiene_left, tiene_right])
    if lados_con_borde >= 3:
        return "cuadro"
    # Subrayado: solo el borde inferior (línea de captura)
    if tiene_bottom and lados_con_borde == 1:
        return "subrayado"
    # Subrayado laxo: bottom presente aunque haya algún lado adicional
    if tiene_bottom:
        return "subrayado"
    return "vacio"


# ──────────────────────────────────────────────────────────────────────────────
# PARSER-02: Detección de líneas de captura divididas entre celdas consecutivas
# ──────────────────────────────────────────────────────────────────────────────

def _calcular_rango_linea_captura(
    hoja: Worksheet,
    fila: int,
    col_inicio: int,
    mapa_merges: Dict[Tuple[int, int], Any],
) -> Tuple[int, int, int]:
    """Escanea dinámicamente hacia la derecha omitiendo celdas espaciadoras sin borde
    para ubicar el inicio exacto (c_inicio) y el fin (c_fin) de una línea de captura continua.

    Returns:
        Tuple[col_inicio_real, col_fin_real, ancho_linea]
    """
    max_col = hoja.max_column or 1
    col_actual = col_inicio
    
    # 1. Saltar celdas espaciadoras vacías sin borde si la línea empieza más a la derecha
    espacios_saltados = 0
    while col_actual <= max_col and espacios_saltados < 4:
        if not _celda_vacia(hoja, fila, col_actual):
            break
        
        bordes = _analizar_bordes_celda(hoja, fila, col_actual)
        rango = mapa_merges.get((fila, col_actual))
        
        if bordes["bottom"] or rango is not None:
            break
            
        col_actual += 1
        espacios_saltados += 1

    if col_actual > max_col or not _celda_vacia(hoja, fila, col_actual):
        col_actual = col_inicio

    c_inicio_real = col_actual
    c_fin_real = col_actual
    ancho = 0

    while col_actual <= max_col:
        if not _celda_vacia(hoja, fila, col_actual):
            break
            
        rango = mapa_merges.get((fila, col_actual))
        if rango is not None:
            c_fin_real = rango.max_col
            ancho += (rango.max_col - col_actual + 1)
            break
            
        bordes = _analizar_bordes_celda(hoja, fila, col_actual)
        if bordes["bottom"]:
            c_fin_real = col_actual
            ancho += 1
            col_actual += 1
        else:
            break

    if ancho == 0:
        return (col_inicio, col_inicio, 1)

    return (c_inicio_real, c_fin_real, max(1, ancho))



# ──────────────────────────────────────────────────────────────────────────────
# PARSER-HYBRID-01: Cálculo determinista de ubicación de escritura (sin LLM)
# ──────────────────────────────────────────────────────────────────────────────

import re as _re

_PATRON_INLINE_GUIONES = _re.compile(r"_{2,}|\.{3,}")


def _calcular_ubicacion_fisica(
    val_rotulo: str,
    derecha_vacia: bool,
    abajo_vacia: bool,
    derecha_es_merge: bool,
    tipo_espacio: str,
    color_fondo: str = "",
) -> str:
    """Calcula la ubicación de escritura usando flags espaciales y estilo visual de la celda.

    Jerarquía determinista:
      1. Si la celda tiene fondo sombreado (color_fondo):
         NUNCA escribir en la misma celda. Redirigir a la celda de entrada con borde/espacio
         a la DERECHA (o ABAJO si es cabecera o derecha bloqueada).
      2. Texto declarativo inline (____) sin fondo sombreado: escribir en la MISMA celda.
      3. Espacio libre a la derecha → DERECHA.
      4. Derecha bloqueada pero abajo libre → ABAJO.
      5. Fallback seguro → DERECHA.
    Returns:
        str: "misma" | "derecha" | "abajo"
    """
    tiene_fondo = bool(color_fondo and str(color_fondo).strip())

    # Regla 0: Celdas con fondo sombreado (gris/color) NUNCA se escriben en 'misma'
    if tiene_fondo:
        espacio_derecha = derecha_vacia or derecha_es_merge
        if espacio_derecha and tipo_espacio != "abajo":
            return "derecha"
        if abajo_vacia:
            return "abajo"
        return "derecha"

    # Regla 1: Texto declarativo con guiones inline sin fondo → escribir MISMA celda
    if _PATRON_INLINE_GUIONES.search(val_rotulo):
        return "misma"

    # Regla 2: Hay espacio libre a la derecha → escribir a la DERECHA
    espacio_derecha_libre = derecha_vacia or derecha_es_merge
    if espacio_derecha_libre and tipo_espacio != "abajo":
        return "derecha"

    # Regla 3: Derecha bloqueada pero abajo libre → cabecera de tabla → ABAJO
    if not espacio_derecha_libre and abajo_vacia:
        return "abajo"

    # Regla 4: Fallback seguro → DERECHA
    return "derecha"


# ──────────────────────────────────────────────────────────────────────────────
# PARSER-HYBRID-02: Escaneo de celdas pre-diligenciadas para reprocesamiento incremental
# ──────────────────────────────────────────────────────────────────────────────

def escanear_celdas_prellenadas(workbook: Any) -> Dict[Tuple[str, int, int], str]:
    """Escanea un workbook cargado y retorna todas las celdas que ya tienen datos escritos.

    Este scanner es la base del Reprocesamiento Incremental: al re-subir un formulario
    parcialmente diligenciado, preserva el 100% de los datos existentes y solo intenta
    completar los campos vacíos restantes.

    Args:
        workbook: Objeto Workbook de openpyxl ya cargado.

    Returns:
        Dict {(nombre_hoja, fila, columna): valor_str} con todas las celdas no vacías.
        Solo incluye celdas que contienen texto real (excluye fórmulas puras y valores None).
    """
    celdas_prellenadas: Dict[Tuple[str, int, int], str] = {}

    for nombre_hoja in workbook.sheetnames:
        ws = workbook[nombre_hoja]
        max_f = ws.max_row or 0
        max_c = ws.max_column or 0

        for fila in range(1, max_f + 1):
            for col in range(1, max_c + 1):
                celda = ws.cell(row=fila, column=col)
                val = celda.value

                # Ignorar celdas vacías o con solo espacios/caracteres invisibles
                if val is None:
                    continue

                txt = str(val).replace("\xa0", " ").replace("\t", " ").strip()
                if not txt or txt.startswith("="):
                    continue

                celdas_prellenadas[(nombre_hoja, fila, col)] = txt

    return celdas_prellenadas


import re


_PATRON_OPCION_CASILLA = re.compile(
    r"^\s*(S[ÍI]|NO|C\.?C\.?|C\.?E\.?|PASAPORTE|NIT|OTRO|AUTORRETENEDOR|GRAN CONTRIBUYENTE|MEDIANA|PEQUEÑA|GRAN EMPRESA|IMPORTADOR|EXPORTADOR|FABRICANTE|DISTRIBUIDOR)\b",
    re.IGNORECASE,
)


def _es_casilla_verificacion(
    vacia: bool,
    es_merge: bool,
    bordes: Dict[str, bool],
    ancho_linea: int,
    texto_rotulo: str = "",
) -> bool:
    """Determina si una celda de destino actúa como casilla de verificación (checkbox 1x1).

    Criterio Híbrido:
      - Condición Física: celda vacía, no es merge, ancho <= 1, lados_con_borde >= 3.
      - Condición Semántica: el rótulo vecino es corto (<= 20 chars) o coincide
        con patrones de opción explícitos (SÍ, NO, CC, CE, etc.).
    """
    if not vacia or es_merge or ancho_linea > 1:
        return False

    lados_con_borde = sum(1 for v in bordes.values() if v)
    if lados_con_borde < 3:
        return False

    if texto_rotulo:
        texto_limpio = str(texto_rotulo).strip()
        if len(texto_limpio) <= 20 or _PATRON_OPCION_CASILLA.search(texto_limpio):
            return True

    return False


# ──────────────────────────────────────────────────────────────────────────────
# PARSER-03: Extracción del color de fondo (fill.fgColor)
# ──────────────────────────────────────────────────────────────────────────────

def _extraer_color_fondo(celda) -> str:
    """Extrae el color de fondo de una celda como string hex (ej. 'FFFF00') o '' si no tiene.

    Soporta colores de tipo 'rgb', 'indexed' y 'theme'.
    Retorna cadena vacía si la celda no tiene relleno significativo (ej. blanco puro o transparente).
    """
    try:
        fill = celda.fill
        if fill is None:
            return ""
        # Solo PatternFill con fill_type definido indica color real
        if not getattr(fill, "fill_type", None) or fill.fill_type in (None, "none"):
            return ""
        fg = getattr(fill, "fgColor", None)
        if fg is None:
            return ""
        color_type = getattr(fg, "type", None)
        if color_type == "rgb":
            try:
                rgb = getattr(fg, "rgb", "") or ""
                if str(rgb) in ("", "00000000", "FFFFFFFF", "00FFFFFF", "FF000000"):
                    return ""
                return str(rgb)[-6:] if len(str(rgb)) == 8 else str(rgb)
            except Exception:
                return ""
        if color_type == "indexed":
            try:
                idx = getattr(fg, "indexed", "")
                if idx in (0, 64, None, ""):
                    return ""
                return f"indexed:{idx}"
            except Exception:
                return ""
        if color_type == "theme":
            try:
                theme = getattr(fg, "theme", None)
                tint = getattr(fg, "tint", 0.0) or 0.0
                if (theme == 0 or theme is None) and abs(tint) < 0.05:
                    return ""
                return f"theme:{theme}:{tint:.2f}"
            except Exception:
                return ""
    except Exception:
        pass
    return ""


# ──────────────────────────────────────────────────────────────────────────────
# PARSER-07: Filtro de hojas no rellenables por nombre
# ──────────────────────────────────────────────────────────────────────────────

# Palabras clave que identifican hojas de instrucciones/portada sin campos rellenables.
# Se compara en minusculas contra el nombre de la hoja.
_NOMBRES_HOJA_EXCLUIR = frozenset({
    "instrucciones", "instruccion", "instruction", "instructions",
    "portada", "cover", "caratula", "carátula",
    "referencias", "reference", "references",
    "glosario", "glossary",
    "terminos", "términos", "terms",
    "condiciones", "conditions",
    "ayuda", "help",
    "leyenda", "legend",
    "resumen", "summary",
    "inicio", "home",
    "indice", "índice", "index",
    "notas", "notes",
})


def _es_hoja_no_rellenable(nombre_hoja: str) -> bool:
    """PARSER-07: Retorna True si la hoja debe ser descartada por su nombre.

    Compara el nombre en minusculas contra la lista de palabras clave excluyentes.
    Tambien descarta hojas cuyo nombre contenga alguna de esas palabras
    (ej. 'Hoja_Instrucciones', 'Portada General').
    """
    nombre_lower = nombre_hoja.strip().lower()
    if nombre_lower in _NOMBRES_HOJA_EXCLUIR:
        return True
    return any(kw in nombre_lower for kw in _NOMBRES_HOJA_EXCLUIR)


# ──────────────────────────────────────────────────────────────────────────────
# PARSER-08: Filtro de ruido en textos de celdas
# ──────────────────────────────────────────────────────────────────────────────

# Patron: fechas en formatos colombianos/internacionales comunes
_PATRON_FECHA = re.compile(
    r'^\d{1,2}[/\-.\s]\d{1,2}[/\-.\s]\d{2,4}$'
    r'|^\d{4}[/\-.]\d{1,2}[/\-.]\d{1,2}$',
    re.IGNORECASE
)
# Patron: numeros puros o con separadores de miles/decimales (no son rotulos)
_PATRON_NUMERO_TEXTO = re.compile(r'^[\d.,\s%$\-+()]{1,20}$')
# Patron: texto compuesto solo de puntuacion, guiones, asteriscos o simbolos
_PATRON_SOLO_PUNTUACION = re.compile(r'^[\W_]+$')
# Patron: codigos CIIU (4 digitos exactos), codigos de actividad o codigos postales cortos
_PATRON_CODIGO_CORTO = re.compile(r'^\d{3,6}$')
# Patron: URLs, correos o cadenas tecnicas que no son rotulos de formulario
_PATRON_URL = re.compile(r'https?://|www\.|@.*\.', re.IGNORECASE)


def _es_texto_ruido(texto: str) -> bool:
    """PARSER-08: Retorna True si el texto de la celda es ruido y debe descartarse.

    Clasifica como ruido:
    - Fechas en cualquier formato comun (dd/mm/yyyy, yyyy-mm-dd, etc.)
    - Numeros puros con separadores de miles o porcentajes
    - Codigos cortos de 3 a 6 digitos (CIIU, codigos postales, etc.)
    - Texto compuesto unicamente de puntuacion o simbolos
    - URLs o correos electronicos
    """
    t = texto.strip()
    if _PATRON_FECHA.match(t):
        return True
    if _PATRON_NUMERO_TEXTO.match(t):
        return True
    if _PATRON_CODIGO_CORTO.match(t):
        return True
    if _PATRON_SOLO_PUNTUACION.match(t):
        return True
    if _PATRON_URL.search(t):
        return True
    return False


def escanear_mapa_formularios(libro) -> List[Dict[str, Any]]:
    """Recorre todas las hojas y extrae el mapa visual/espacial de los rótulos con texto.

    Optimizado con indexación de rangos combinados O(1) y escaneo de baja latencia.
    """
    formulario: List[Dict[str, Any]] = []
    for hoja in libro.worksheets:
        # PARSER-07: Saltar hojas de instrucciones, portada, glosario y similares
        if _es_hoja_no_rellenable(hoja.title):
            print(f"[AutoForm AI PARSER-07] Hoja '{hoja.title}' omitida (tipo no rellenable).")
            continue
        # 1. Crear un diccionario O(1) de celdas combinadas -> objeto CellRange
        mapa_merges: Dict[Tuple[int, int], Any] = {}
        for rango in hoja.merged_cells.ranges:
            for r in range(rango.min_row, rango.max_row + 1):
                for c in range(rango.min_col, rango.max_col + 1):
                    mapa_merges[(r, c)] = rango

        # 2. Recorrer únicamente las celdas que realmente contienen datos
        for row in hoja.iter_rows():
            for cell in row:
                valor = cell.value
                if valor is None:
                    continue

                # Omitir tipos de datos numéricos o booleanos (no son etiquetas de formulario)
                if isinstance(valor, (int, float, bool)):
                    continue

                texto = str(valor).strip()
                # Omitir textos vacíos o demasiado cortos (menor a 2)
                # FIX-2: Ampliado límite superior a 200 para capturar rótulos compuestos largos
                if not texto or len(texto) < 2 or len(texto) > 200:
                    continue

                # PARSER-08: Filtrar ruido: fechas, numeros, codigos, puntuacion y URLs
                if _es_texto_ruido(texto):
                    continue

                fila = cell.row
                columna = cell.column

                # ── PARSER-05: Merge propio de la etiqueta ─────────────────────
                merge_propio = mapa_merges.get((fila, columna))
                if merge_propio is not None:
                    es_celda_principal = (
                        merge_propio.min_row == fila and merge_propio.min_col == columna
                    )
                    if not es_celda_principal:
                        # Sub-celda de un merge ya registrado → saltar para evitar duplicados
                        continue
                    derecha_fila    = fila
                    derecha_columna = merge_propio.max_col + 1
                    es_merge_principal = True
                    coord_merge = str(merge_propio)
                else:
                    derecha_fila    = fila
                    derecha_columna = columna + 1
                    es_merge_principal = False
                    coord_merge = ""
                # ────────────────────────────────────────────────────────────

                abajo_fila   = fila + 1
                abajo_columna = columna

                # 3. Estado de vecinos: vacío / combinado (consultas O(1))
                max_hoja_col = hoja.max_column or 1
                derecha_vacia   = _celda_vacia(hoja, derecha_fila, derecha_columna) if derecha_columna <= max_hoja_col else False
                abajo_vacia     = _celda_vacia(hoja, abajo_fila, abajo_columna)

                rango_derecha   = mapa_merges.get((derecha_fila, derecha_columna))
                rango_abajo     = mapa_merges.get((abajo_fila, abajo_columna))
                derecha_es_merge = rango_derecha is not None
                abajo_es_merge   = rango_abajo is not None

                # FIX-1: Considerar "placeholder" como vacío efectivo.
                # Celdas con solo guiones, puntos o espacios (p.ej. "____", "...", "-")
                # no son contenido real — el rótulo no debe descartarse por ellas.
                def _es_placeholder_celda(h: Worksheet, f: int, c: int) -> bool:
                    if f < 1 or c < 1 or f > (h.max_row or 0) or c > (h.max_column or 0):
                        return False
                    v = h.cell(row=f, column=c).value
                    if v is None:
                        return False
                    return bool(_re.match(r'^[\s_\.\:\-]+$', str(v).strip()))

                derecha_placeholder = _es_placeholder_celda(hoja, derecha_fila, derecha_columna)
                abajo_placeholder   = _es_placeholder_celda(hoja, abajo_fila, abajo_columna)

                # Actualizar flags: si el vecino es placeholder, se trata como vacío
                if derecha_placeholder:
                    derecha_vacia = True
                if abajo_placeholder:
                    abajo_vacia = True

                # ── Análisis de celda superior (arriba) para líneas de firma y encabezados inferiores ──
                arriba_fila = fila - 1
                arriba_columna = columna
                arriba_vacia = False
                arriba_con_borde_inferior = False
                celda_con_borde_superior = False

                if arriba_fila >= 1:
                    arriba_vacia = _celda_vacia(hoja, arriba_fila, arriba_columna) or _es_placeholder_celda(hoja, arriba_fila, arriba_columna)
                    bordes_arriba = _analizar_bordes_celda(hoja, arriba_fila, arriba_columna)
                    arriba_con_borde_inferior = bordes_arriba["bottom"]

                bordes_propios = _analizar_bordes_celda(hoja, fila, columna)
                celda_con_borde_superior = bordes_propios["top"]

                # Solo aplicar arriba si la celda es semánticamente un rótulo de firma/firmante/identificación
                es_rotulo_firma = bool(_PATRON_OPCION_CASILLA.search(texto) or re.search(
                    r"\b(?:firma|firmante|nombres?\s+y\s+apellidos?|nombres?|apellidos?|representante(?:\s+legal)?|revisor(?:\s+fiscal)?|contador|apoderado|huella|c\.?c\.?|c[eé]dula|identificaci[oó]n|n[oú]mero\s+de\s+identificaci[oó]n|cargo|elabor[oó]|revis[oó]|aprob[oó]|solicitante|titular)\b",
                    texto,
                    re.IGNORECASE
                ))

                tiene_linea_firma_arriba = bool(
                    arriba_vacia and (celda_con_borde_superior or arriba_con_borde_inferior) and es_rotulo_firma
                )

                # Optimización: si ningún vecino tiene espacio ni es merge ni tiene línea superior de firma, descartar.
                if not derecha_vacia and not abajo_vacia and not derecha_es_merge and not abajo_es_merge and not tiene_linea_firma_arriba:
                    continue

                # ── PARSER-01: Análisis de bordes visuales ──────────────────
                bordes_derecha = _analizar_bordes_celda(hoja, derecha_fila, derecha_columna)
                bordes_abajo   = _analizar_bordes_celda(hoja, abajo_fila, abajo_columna)

                derecha_con_borde_inferior = bordes_derecha["bottom"]
                derecha_con_borde_todo     = all(bordes_derecha.values())
                abajo_con_borde_inferior   = bordes_abajo["bottom"]
                abajo_con_borde_todo       = all(bordes_abajo.values())

                tipo_derecha = _calcular_tipo_espacio(derecha_vacia, derecha_es_merge, bordes_derecha)
                tipo_abajo   = _calcular_tipo_espacio(abajo_vacia,   abajo_es_merge,   bordes_abajo)

                # El tipoEspacioEscritura representa el espacio preferido:
                # Si tiene línea de firma arriba y derecha no es un cuadro/subrayado explícito, preferir arriba
                if tiene_linea_firma_arriba and tipo_derecha not in ("merge", "subrayado", "cuadro"):
                    tipo_espacio_escritura = "arriba"
                elif tipo_derecha in ("merge", "subrayado", "cuadro", "vacio"):
                    tipo_espacio_escritura = tipo_derecha
                else:
                    tipo_espacio_escritura = tipo_abajo

                # ── PARSER-02: Rango dinámico y ancho real de la línea de captura ────
                c_inicio_linea, c_fin_linea, ancho_linea = _calcular_rango_linea_captura(
                    hoja, fila, derecha_columna, mapa_merges
                )

                # ── PARSER-04: Ancho del rango combinado del vecino derecho ───
                if rango_derecha is not None:
                    ancho_merge_vecino = rango_derecha.max_col - rango_derecha.min_col + 1
                else:
                    ancho_merge_vecino = 1
                # ────────────────────────────────────────────────────────────

                # ── Detección Híbrida de Casillas de Verificación (Checkbox 1x1) ──
                es_casilla = _es_casilla_verificacion(
                    derecha_vacia, derecha_es_merge, bordes_derecha, ancho_linea, texto
                ) or _es_casilla_verificacion(
                    abajo_vacia, abajo_es_merge, bordes_abajo, 1, texto
                )

                # ── PARSER-03: Color de fondo del rótulo ─────────────────────
                color_fondo = _extraer_color_fondo(cell)
                # ────────────────────────────────────────────────────────────

                formulario.append(
                    {
                        "hoja": hoja.title,
                        "fila": fila,
                        "columna": columna,
                        "valor": texto,
                        "derechaVacia":   derecha_vacia,
                        "abajoVacia":     abajo_vacia,
                        "arribaVacia":    arriba_vacia,
                        "tieneLineaFirmaArriba": tiene_linea_firma_arriba,
                        "celdaConBordeSuperior": celda_con_borde_superior,
                        "arribaConBordeInferior": arriba_con_borde_inferior,
                        "derechaEsMerge": derecha_es_merge,
                        "abajoEsMerge":   abajo_es_merge,
                        "derechaConBordeInferior": derecha_con_borde_inferior,
                        "derechaConBordeTodo":     derecha_con_borde_todo,
                        "abajoConBordeInferior":   abajo_con_borde_inferior,
                        "abajoConBordeTodo":       abajo_con_borde_todo,
                        "tipoEspacioEscritura": tipo_espacio_escritura,
                        "inicioLineaCol": c_inicio_linea,
                        "finLineaCol":    c_fin_linea,
                        "anchoLinea":     ancho_linea,
                        "anchoMergeVecino": ancho_merge_vecino,
                        "esMergePrincipal": es_merge_principal,
                        "coordMerge":       coord_merge,
                        "esCasillaVerificacion": es_casilla,
                        "colorFondo": color_fondo,
                    }
                )
    return formulario


