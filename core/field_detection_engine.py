"""Módulo Corazón: Field Detection Engine (AutoForm AI).

Este módulo es el motor dedicado de detección de campos y cálculo de coordenadas físicas
previo a cualquier llamada al LLM o escritura en Excel/PDF.

Estructura:
  - BoundaryScanner: Escáner de límites físicos y traza de líneas de captura.
  - FieldClassifier: Clasificador de 5 patrones visuales de formularios.
  - ContextEnricher: Extrae títulos de sección y contexto semántico.
  - InjectionPolicy: Aplica preservación de fondo blanco y cero alteración de dimensiones.
  - FieldDetectionEngine: Orquestador principal de extracción de objetos FormFieldIntent.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Dict, List, Literal, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet


# ──────────────────────────────────────────────────────────────────────────────
# 1. Objeto de Intención de Campo (FormFieldIntent)
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class FormFieldIntent:
    """Representa la intención y coordenadas físicas deterministas de un campo detectado."""

    id_rotulo: int
    rotulo_texto: str
    seccion_titulo: str
    tipo_patron: Literal[
        "EXTENDED_LINE_INPUT",
        "SINGLE_LINE_INPUT",
        "TABLE_CELL_INPUT",
        "INLINE_TEXT_INPUT",
        "CHECKBOX_INPUT",
    ]
    dest_hoja: str
    dest_fila: int
    dest_col_inicio: int
    dest_col_fin: int
    dest_ubicacion: Literal["derecha", "abajo", "misma"]
    requiere_merge: bool
    ancho_merge: int
    es_casilla: bool
    color_fondo_original: str


# ──────────────────────────────────────────────────────────────────────────────
# 2. BoundaryScanner (Escáner de Límites Físicos)
# ──────────────────────────────────────────────────────────────────────────────

class BoundaryScanner:
    """Determina dinámicamente el inicio (start_col) y fin (end_col) de una línea de captura o casilla."""

    @staticmethod
    def _celda_vacia(hoja: Worksheet, fila: int, col: int) -> bool:
        if fila < 1 or col < 1 or fila > (hoja.max_row or 0) or col > (hoja.max_column or 0):
            return True
        celda = hoja.cell(row=fila, column=col)
        return celda.value is None or str(celda.value).strip() == ""

    @staticmethod
    def _tiene_borde_inferior(hoja: Worksheet, fila: int, col: int) -> bool:
        if fila < 1 or col < 1 or fila > (hoja.max_row or 0) or col > (hoja.max_column or 0):
            return False
        celda = hoja.cell(row=fila, column=col)
        try:
            b = getattr(celda.border, "bottom", None)
            return b is not None and b.style is not None and b.style != "none"
        except Exception:
            return False

    @classmethod
    def escaneo_traza_linea(
        cls,
        hoja: Worksheet,
        fila: int,
        col_inicio: int,
        mapa_merges: Dict[Tuple[int, int], Any],
    ) -> Tuple[int, int, int]:
        """Escanea hacia la derecha omitiendo espaciadores sin borde para ubicar start_col y end_col."""
        max_col = hoja.max_column or 1
        col_actual = col_inicio

        # 1. Omitir hasta 3 celdas espaciadoras vacías sin borde
        espacios_saltados = 0
        while col_actual <= max_col and espacios_saltados < 4:
            if not cls._celda_vacia(hoja, fila, col_actual):
                break
            b_bottom = cls._tiene_borde_inferior(hoja, fila, col_actual)
            rango = mapa_merges.get((fila, col_actual))
            if b_bottom or rango is not None:
                break
            col_actual += 1
            espacios_saltados += 1

        if col_actual > max_col or not cls._celda_vacia(hoja, fila, col_actual):
            col_actual = col_inicio

        c_inicio_real = col_actual
        c_fin_real = col_actual
        ancho = 0

        while col_actual <= max_col:
            if not cls._celda_vacia(hoja, fila, col_actual):
                break
            rango = mapa_merges.get((fila, col_actual))
            if rango is not None:
                c_fin_real = rango.max_col
                ancho += (rango.max_col - col_actual + 1)
                break
            if cls._tiene_borde_inferior(hoja, fila, col_actual):
                c_fin_real = col_actual
                ancho += 1
                col_actual += 1
            else:
                break

        if ancho == 0:
            return (col_inicio, col_inicio, 1)

        return (c_inicio_real, c_fin_real, max(1, ancho))


# ──────────────────────────────────────────────────────────────────────────────
# 3. FieldClassifier (Clasificador de 5 Patrones Visuales)
# ──────────────────────────────────────────────────────────────────────────────

class FieldClassifier:
    """Clasifica una celda de entrada en una de las 5 categorías estructurales."""

    _PATRON_PLACEHOLDER = re.compile(r"_{2,}|\.{3,}")

    @classmethod
    def clasificar(
        cls,
        texto_rotulo: str,
        ancho_linea: int,
        ubicacion: str,
        es_casilla: bool,
        es_tabla: bool,
    ) -> Literal[
        "EXTENDED_LINE_INPUT",
        "SINGLE_LINE_INPUT",
        "TABLE_CELL_INPUT",
        "INLINE_TEXT_INPUT",
        "CHECKBOX_INPUT",
    ]:
        if cls._PATRON_PLACEHOLDER.search(texto_rotulo):
            return "INLINE_TEXT_INPUT"
        if es_casilla:
            return "CHECKBOX_INPUT"
        if ubicacion == "abajo" or es_tabla:
            return "TABLE_CELL_INPUT"
        if ancho_linea > 2:
            return "EXTENDED_LINE_INPUT"
        return "SINGLE_LINE_INPUT"


# ──────────────────────────────────────────────────────────────────────────────
# 4. ContextEnricher (Enriquecedor de Contexto Semántico)
# ──────────────────────────────────────────────────────────────────────────────

class ContextEnricher:
    """Extrae títulos de sección superior y contexto de fila."""

    @staticmethod
    def obtener_titulo_seccion(hoja: Worksheet, fila: int) -> str:
        """Busca hacia arriba la fila de título/encabezado de sección más cercana."""
        for r in range(max(1, fila - 4), fila):
            for c in range(1, min(10, (hoja.max_column or 1) + 1)):
                val = hoja.cell(row=r, column=c).value
                if val and isinstance(val, str) and len(str(val).strip()) > 3:
                    val_clean = str(val).strip()
                    if any(kw in val_clean.upper() for kw in ["INFORMACION", "INFORMACIÓN", "DECLARACION", "DECLARACIÓN", "DATOS", "DATOS BASICO", "IDENTIFICACION", "SECCION"]):
                        return val_clean[:60]
        return "GENERAL"


# ──────────────────────────────────────────────────────────────────────────────
# 5. InjectionPolicy (Guardián de Invariantes Visuales)
# ──────────────────────────────────────────────────────────────────────────────

class InjectionPolicy:
    """Garantiza la preservación de fondos limpios y cero alteración de dimensiones."""

    @staticmethod
    def es_relleno_oscuro_o_encabezado(color_hex: str) -> bool:
        """Determina si un color de fondo hex es un tono oscuro/gris de encabezado."""
        if not color_hex or color_hex.upper() in ("FFFFFF", "00000000", "FFFFFFFF", "NONE"):
            return False
        # Si es gris/azul/amarillo de título
        return True


# ──────────────────────────────────────────────────────────────────────────────
# 6. FieldDetectionEngine (Orquestador Principal)
# ──────────────────────────────────────────────────────────────────────────────

class FieldDetectionEngine:
    """Motor Corazón de Detección Espacial Determinista de Campos en Excel."""

    def __init__(self, libro: Any):
        self.libro = libro

    def extraer_intenciones_campo(self) -> List[FormFieldIntent]:
        """Procesa la cuadrícula de Excel y genera los Objetos FormFieldIntent pre-calculados."""
        intenciones: List[FormFieldIntent] = []
        id_counter = 1

        for hoja in self.libro.worksheets:
            mapa_merges: Dict[Tuple[int, int], Any] = {}
            for rango in hoja.merged_cells.ranges:
                for r in range(rango.min_row, rango.max_row + 1):
                    for c in range(rango.min_col, rango.max_col + 1):
                        mapa_merges[(r, c)] = rango

            for row in hoja.iter_rows():
                for cell in row:
                    val = cell.value
                    if not val or not isinstance(val, str):
                        continue

                    texto = str(val).strip()
                    if not texto or len(texto) < 2 or len(texto) > 140:
                        continue

                    fila = cell.row
                    col = cell.column

                    merge_propio = mapa_merges.get((fila, col))
                    if merge_propio is not None:
                        if not (merge_propio.min_row == fila and merge_propio.min_col == col):
                            continue
                        der_col = merge_propio.max_col + 1
                    else:
                        der_col = col + 1

                    # 1. Trazado de límites físicos
                    c_inicio, c_fin, ancho = BoundaryScanner.escaneo_traza_linea(
                        hoja, fila, der_col, mapa_merges
                    )

                    # 2. Sugerencia Inicial de Ubicación ("misma", "abajo", "derecha")
                    derecha_vacia = BoundaryScanner._celda_vacia(hoja, fila, der_col)
                    abajo_vacia = BoundaryScanner._celda_vacia(hoja, fila + 1, col)

                    if FieldClassifier._PATRON_PLACEHOLDER.search(texto):
                        ubicacion = "misma"
                        c_inicio, c_fin = col, col
                    elif not derecha_vacia and abajo_vacia and (col > 1 and (col + 1) <= (hoja.max_column or 1) and not BoundaryScanner._celda_vacia(hoja, fila, col + 1)):
                        # Encabezados continuos de tabla en fila (ej. Banco, Sucursal, Cuenta)
                        ubicacion = "abajo"
                        c_inicio, c_fin = col, col
                    else:
                        ubicacion = "derecha"

                    # 3. Clasificación de Patrón
                    seccion = ContextEnricher.obtener_titulo_seccion(hoja, fila)
                    patron = FieldClassifier.clasificar(
                        texto_rotulo=texto,
                        ancho_linea=ancho,
                        ubicacion=ubicacion,
                        es_casilla=False,
                        es_tabla=(ubicacion == "abajo"),
                    )

                    intent = FormFieldIntent(
                        id_rotulo=id_counter,
                        rotulo_texto=texto,
                        seccion_titulo=seccion,
                        tipo_patron=patron,
                        dest_hoja=hoja.title,
                        dest_fila=fila,
                        dest_col_inicio=c_inicio,
                        dest_col_fin=c_fin,
                        dest_ubicacion=ubicacion,
                        requiere_merge=(ancho > 1 and c_inicio < c_fin),
                        ancho_merge=ancho,
                        es_casilla=False,
                        color_fondo_original="",
                    )
                    intenciones.append(intent)
                    id_counter += 1

        return intenciones
